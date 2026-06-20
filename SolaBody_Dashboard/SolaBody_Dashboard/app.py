"""
Sola Body — Daily Revenue Command Center
Flask Backend
Run: python app.py  →  open http://localhost:5000
"""

import io
import uuid
import warnings
from datetime import datetime, timedelta

import numpy as np
import pandas as pd
from flask import Flask, jsonify, render_template, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

app = Flask(__name__)
app.secret_key = "sola-dashboard-2026"

# ── In-memory session store (session_id → DataFrame) ──────────
DATA_STORE: dict[str, pd.DataFrame] = {}

# ── Constants ─────────────────────────────────────────────────
EXCEL_ORIGIN   = datetime(1899, 12, 30)
PLATFORM_COLS  = ["shopify", "shopee", "lazada", "tiktok"]
AD_COLS        = ["meta_ads", "shopee_ads", "lazada_ads",
                  "tiktok_ads", "google_ads", "shopee_cpas", "lazada_cpas"]
COL_MAP        = {
    0: "dow", 1: "month", 2: "date_raw",
    3: "shopify", 4: "shopee", 5: "lazada", 6: "tiktok",
    7: "total_revenue", 8: "total_revenue_ex_vat",
    9: "meta_ads", 10: "shopee_ads", 11: "lazada_ads",
    12: "tiktok_ads", 13: "google_ads",
    14: "shopee_cpas", 15: "lazada_cpas",
    20: "total_ads_spent",
}


# ─────────────────────────────────────────────────────────────
#  DATA PROCESSING
# ─────────────────────────────────────────────────────────────

def _to_float(val) -> float:
    if pd.isna(val):
        return 0.0
    if isinstance(val, (int, float)):
        return 0.0 if (isinstance(val, float) and np.isnan(val)) else float(val)
    s = str(val).replace("₱", "").replace(",", "").replace(" ", "").replace("-", "").strip()
    try:
        return float(s) if s else 0.0
    except ValueError:
        return 0.0


def _parse_date(val):
    if pd.isna(val) or str(val).strip() in ("", "nan"):
        return None
    try:
        parsed = pd.to_datetime(val)
        if pd.notna(parsed):
            return parsed
    except Exception:
        pass
    try:
        n = float(val)
        if 40000 < n < 60000:
            return pd.Timestamp(EXCEL_ORIGIN + timedelta(days=int(n)))
    except Exception:
        pass
    return None


def load_and_clean(file) -> pd.DataFrame:
    """
    Read BR Input Excel and return a clean daily DataFrame.
    Handles: two-row headers, currency strings, datetime/serial dates,
    month-total rows, and inconsistent first column.
    """
    xl     = pd.ExcelFile(file)
    sheet  = next(
        (s for s in xl.sheet_names if "daily" in s.lower()),
        xl.sheet_names[0],
    )
    raw    = pd.read_excel(file, sheet_name=sheet, header=None, dtype=str)

    # Find header row (contains "DATE" or "Shopify")
    header_idx = 0
    for i, row in raw.iterrows():
        vals = [str(v).strip().lower() for v in row if pd.notna(v) and str(v).strip()]
        if "date" in vals or "shopify" in vals:
            header_idx = i
            break

    # Data starts 3 rows after header (sub-header + 2 blank rows)
    data = raw.iloc[header_idx + 3 :].copy().reset_index(drop=True)
    data.columns = [COL_MAP.get(i, f"col_{i}") for i in range(len(data.columns))]

    # Parse dates; drop invalid rows (totals, blanks)
    data["date"] = pd.to_datetime(
        data["date_raw"].apply(_parse_date), errors="coerce"
    )
    df = data[data["date"].notna()].copy()

    # Clean numeric columns
    num_cols = PLATFORM_COLS + ["total_revenue", "total_revenue_ex_vat"] + AD_COLS + ["total_ads_spent"]
    for col in num_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col].apply(_to_float), errors="coerce").fillna(0.0)

    # Derived columns
    rev   = df["total_revenue"].astype(float)
    spend = df["total_ads_spent"].astype(float)
    df["roas"]        = np.where(spend > 0, rev / spend, 0.0)
    df["week_number"] = df["date"].dt.isocalendar().week.astype(int)
    df["month_name"]  = df["date"].dt.strftime("%b %Y")
    df["day_of_week"] = df["date"].dt.day_name()
    df["remarks"]     = df["dow"].apply(
        lambda v: str(v).strip()
        if pd.notna(v)
        and len(str(v).strip()) > 3
        and str(v).strip().upper()
        not in {"SUN", "MON", "TUE", "WED", "THU", "FRI", "SAT", "NAN"}
        else ""
    )

    return df.sort_values("date").reset_index(drop=True)


def _pct(new_val, old_df, col) -> float | None:
    if old_df.empty:
        return None
    old = float(old_df.iloc[0][col])
    if old == 0:
        return None
    return round((new_val - old) / abs(old) * 100, 1)


# ─────────────────────────────────────────────────────────────
#  EXCEL REPORT HELPERS
# ─────────────────────────────────────────────────────────────

DARK   = "1D3557"
MID    = "457B9D"
LIGHT  = "A8DADC"
ALT    = "F1FAEE"
THIN   = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(color="FFFFFF", size=10, bold=True) -> Font:
    return Font(name="Arial", bold=bold, color=color, size=size)

def _write_headers(ws, title: str, cols: list, row: int = 1):
    n = len(cols)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font      = _font(size=12)
    cell.fill      = _fill(DARK)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 22
    for j, name in enumerate(cols, 1):
        c = ws.cell(row=row + 1, column=j, value=name)
        c.font      = _font(size=10)
        c.fill      = _fill(MID)
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border    = THIN
    ws.row_dimensions[row + 1].height = 28

def _autofit(ws, min_w=10, max_w=30):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width  = max(min_w, min(max_w, max(len(str(c.value or "")) for c in col) + 2))
        ws.column_dimensions[letter].width = width

def _style(ws, row: int, n_cols: int, alt: bool,
           currency=(), pct=(), roas=()):
    fill = _fill(ALT) if alt else None
    for j in range(1, n_cols + 1):
        c = ws.cell(row=row, column=j)
        c.border = THIN
        c.font   = Font(name="Arial", size=10)
        if fill:
            c.fill = fill
        if j in currency:
            c.number_format = "₱#,##0.00"
        if j in pct and isinstance(c.value, float):
            c.number_format = "0.0%"
            c.font = Font(name="Arial", size=10, bold=True,
                          color="006400" if c.value >= 0 else "CC0000")
        if j in roas and isinstance(c.value, (int, float)):
            c.number_format = "0.00"
            color = "006400" if c.value >= 4 else ("CC6600" if c.value >= 2 else "CC0000")
            c.font = Font(name="Arial", size=10, bold=True, color=color)

def _total(ws, row: int, values: list, currency=()):
    for j, v in enumerate(values, 1):
        c = ws.cell(row=row, column=j, value=v)
        c.fill   = _fill(LIGHT)
        c.font   = Font(name="Arial", size=10, bold=True)
        c.border = THIN
        if j in currency:
            c.number_format = "₱#,##0.00"

def _pdelta(new_val, old_val):
    try:
        if old_val and old_val != 0:
            return (new_val - old_val) / abs(old_val)
    except Exception:
        pass
    return None


# ─────────────────────────────────────────────────────────────
#  REPORT 1 — DAILY
# ─────────────────────────────────────────────────────────────

def generate_daily_report(df: pd.DataFrame, sel: pd.Timestamp) -> io.BytesIO | None:
    row  = df[df["date"] == sel]
    if row.empty:
        return None
    t    = row.iloc[0]
    y    = df[df["date"] == sel - pd.Timedelta(days=1)]
    lw   = df[df["date"] == sel - pd.Timedelta(days=7)]
    mtd  = df[(df["date"].dt.month == sel.month) & (df["date"].dt.year == sel.year) & (df["date"] <= sel)]

    wb = Workbook()

    # Sheet 1 — Snapshot
    ws1 = wb.active
    ws1.title = "1. Snapshot"
    _write_headers(ws1, f"Daily Snapshot — {sel.strftime('%B %d, %Y')}",
                   ["Metric", "Value", "vs Yesterday", "vs Last Week"])

    yrev  = float(y.iloc[0]["total_revenue"])  if not y.empty  else None
    lwrev = float(lw.iloc[0]["total_revenue"]) if not lw.empty else None
    yads  = float(y.iloc[0]["total_ads_spent"]) if not y.empty else None
    mtdr  = float(mtd["total_revenue"].sum())
    mtda  = float(mtd["total_ads_spent"].sum())

    rows1 = [
        ("Total Revenue",        float(t["total_revenue"]),    _pdelta(float(t["total_revenue"]),    yrev),  _pdelta(float(t["total_revenue"]),  lwrev)),
        ("Total Revenue ex-VAT", float(t["total_revenue_ex_vat"]), None, None),
        ("Total Ad Spend",       float(t["total_ads_spent"]),  _pdelta(float(t["total_ads_spent"]),  yads),  None),
        ("ROAS",                 round(float(t["roas"]), 2),   None, None),
        ("MTD Revenue",          mtdr, None, None),
        ("MTD Ad Spend",         mtda, None, None),
        ("MTD ROAS",             round(mtdr / mtda, 2) if mtda > 0 else 0, None, None),
        ("Campaign / Event",     str(t.get("remarks") or "—"), None, None),
    ]
    for i, (label, val, d1, d2) in enumerate(rows1, 3):
        ws1.cell(row=i, column=1, value=label)
        ws1.cell(row=i, column=2, value=val)
        ws1.cell(row=i, column=3, value=d1)
        ws1.cell(row=i, column=4, value=d2)
        is_cur = label not in ("ROAS", "MTD ROAS", "Campaign / Event")
        _style(ws1, i, 4, i % 2 == 0, currency=(2,) if is_cur else (), pct=(3, 4))
        if "ROAS" in label:
            ws1.cell(row=i, column=2).number_format = "0.00"
    _autofit(ws1)

    # Sheet 2 — Platform Revenue
    ws2 = wb.create_sheet("2. Platform Revenue")
    _write_headers(ws2, f"Platform Revenue — {sel.strftime('%B %d, %Y')}",
                   ["Platform", "Revenue", "% of Total", "vs Yesterday", "vs Last Week"])
    total = float(t["total_revenue"])
    for i, (name, col) in enumerate([("Shopify","shopify"),("Shopee","shopee"),
                                       ("Lazada","lazada"),("TikTok Shop","tiktok")], 3):
        val   = float(t[col])
        share = val / total if total > 0 else 0
        vy    = _pdelta(val, float(y.iloc[0][col])  if not y.empty  else None)
        vlw   = _pdelta(val, float(lw.iloc[0][col]) if not lw.empty else None)
        ws2.cell(row=i, column=1, value=name)
        ws2.cell(row=i, column=2, value=val)
        ws2.cell(row=i, column=3, value=share)
        ws2.cell(row=i, column=4, value=vy)
        ws2.cell(row=i, column=5, value=vlw)
        _style(ws2, i, 5, i % 2 == 0, currency=(2,), pct=(3, 4, 5))
    last2 = 4 + 2
    _total(ws2, last2 + 1, ["TOTAL", None, None, None, None], currency=(2,))
    ws2.cell(row=last2+1, column=2).value = f"=SUM(B3:B{last2})"
    ws2.cell(row=last2+1, column=2).number_format = "₱#,##0.00"
    _autofit(ws2)

    # Sheet 3 — Ad Spend
    ws3 = wb.create_sheet("3. Ad Spend")
    _write_headers(ws3, f"Ad Spend — {sel.strftime('%B %d, %Y')}",
                   ["Channel", "Spend", "% of Budget", "Platform Revenue", "ROAS"])
    channels = [("Meta Ads","meta_ads","shopify"),("Shopee Ads","shopee_ads","shopee"),
                ("Lazada Ads","lazada_ads","lazada"),("TikTok Ads","tiktok_ads","tiktok"),
                ("Google Ads","google_ads","shopify"),("Shopee CPAs","shopee_cpas","shopee"),
                ("Lazada CPAs","lazada_cpas","lazada")]
    total_spend = float(t["total_ads_spent"])
    for i, (name, sc, rc) in enumerate(channels, 3):
        spend = float(t[sc]); rev = float(t[rc])
        share = spend / total_spend if total_spend > 0 else 0
        roas  = round(rev / spend, 2) if spend > 0 else 0
        ws3.cell(row=i, column=1, value=name)
        ws3.cell(row=i, column=2, value=spend)
        ws3.cell(row=i, column=3, value=share)
        ws3.cell(row=i, column=4, value=rev)
        ws3.cell(row=i, column=5, value=roas)
        _style(ws3, i, 5, i % 2 == 0, currency=(2, 4), pct=(3,), roas=(5,))
    last3 = len(channels) + 2
    _total(ws3, last3+1, ["TOTAL", None, None, None, None], currency=(2, 4))
    ws3.cell(row=last3+1, column=2).value = f"=SUM(B3:B{last3})"
    ws3.cell(row=last3+1, column=2).number_format = "₱#,##0.00"
    ws3.cell(row=last3+1, column=4).value = total
    ws3.cell(row=last3+1, column=4).number_format = "₱#,##0.00"
    ws3.cell(row=last3+1, column=5).value = round(float(t["roas"]), 2)
    ws3.cell(row=last3+1, column=5).number_format = "0.00"
    _autofit(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────
#  REPORT 2 — WEEKLY
# ─────────────────────────────────────────────────────────────

def generate_weekly_report(df: pd.DataFrame) -> io.BytesIO | None:
    last  = df["date"].max()
    wk_no = int(last.isocalendar().week)
    yr    = last.year
    wk    = df[(df["week_number"] == wk_no) & (df["date"].dt.year == yr)]
    prev  = df[(df["week_number"] == wk_no - 1) & (df["date"].dt.year == yr)]
    if wk.empty:
        return None

    wb   = Workbook()
    ws   = wk["date"].min().strftime("%b %d")
    we   = wk["date"].max().strftime("%b %d, %Y")

    # Sheet 1 — Summary
    ws1  = wb.active
    ws1.title = "1. Weekly Summary"
    _write_headers(ws1, f"Weekly Summary — {ws} to {we}",
                   ["Metric", "This Week", "Last Week", "WoW Change"])
    this_rev   = float(wk["total_revenue"].sum())
    prev_rev   = float(prev["total_revenue"].sum()) if not prev.empty else 0
    this_spend = float(wk["total_ads_spent"].sum())
    prev_spend = float(prev["total_ads_spent"].sum()) if not prev.empty else 0
    pw = lambda col: float(prev[col].sum()) if not prev.empty else 0

    summary = [
        ("Total Revenue",  this_rev,              prev_rev,     _pdelta(this_rev,  prev_rev)),
        ("Shopify",        float(wk["shopify"].sum()), pw("shopify"), _pdelta(float(wk["shopify"].sum()), pw("shopify"))),
        ("Shopee",         float(wk["shopee"].sum()),  pw("shopee"),  _pdelta(float(wk["shopee"].sum()),  pw("shopee"))),
        ("Lazada",         float(wk["lazada"].sum()),  pw("lazada"),  _pdelta(float(wk["lazada"].sum()),  pw("lazada"))),
        ("TikTok Shop",    float(wk["tiktok"].sum()),  pw("tiktok"),  _pdelta(float(wk["tiktok"].sum()),  pw("tiktok"))),
        ("Total Ad Spend", this_spend,            prev_spend,   _pdelta(this_spend, prev_spend)),
        ("ROAS",           round(this_rev/this_spend,2) if this_spend else 0,
                           round(prev_rev/prev_spend,2) if prev_spend else 0, None),
        ("Best Day",  wk.loc[wk["total_revenue"].idxmax(), "day_of_week"], "—", None),
        ("Worst Day", wk.loc[wk["total_revenue"].idxmin(), "day_of_week"], "—", None),
    ]
    for i, (label, tw, pw_val, wow) in enumerate(summary, 3):
        ws1.cell(row=i, column=1, value=label)
        ws1.cell(row=i, column=2, value=tw)
        ws1.cell(row=i, column=3, value=pw_val)
        ws1.cell(row=i, column=4, value=wow)
        is_cur = label not in ("ROAS", "Best Day", "Worst Day")
        _style(ws1, i, 4, i % 2 == 0, currency=(2, 3) if is_cur else (), pct=(4,))
        if label == "ROAS":
            ws1.cell(row=i, column=2).number_format = "0.00"
            ws1.cell(row=i, column=3).number_format = "0.00"
    _autofit(ws1)

    # Sheet 2 — Daily Breakdown
    ws2 = wb.create_sheet("2. Daily Breakdown")
    _write_headers(ws2, f"Daily Breakdown — Week {wk_no}",
                   ["Date", "Day", "Shopify", "Shopee", "Lazada", "TikTok", "Total Revenue", "Ad Spend", "ROAS", "Campaign"])
    for i, (_, r) in enumerate(wk.iterrows(), 3):
        vals = [r["date"].strftime("%Y-%m-%d"), r["day_of_week"],
                float(r["shopify"]), float(r["shopee"]), float(r["lazada"]), float(r["tiktok"]),
                float(r["total_revenue"]), float(r["total_ads_spent"]),
                round(float(r["roas"]), 2), str(r.get("remarks") or "")]
        for j, v in enumerate(vals, 1):
            ws2.cell(row=i, column=j, value=v)
        _style(ws2, i, 10, i % 2 == 0, currency=(3,4,5,6,7,8), roas=(9,))
    last2 = len(wk) + 2
    _total(ws2, last2+1, ["TOTAL",""] + [None]*8, currency=(3,4,5,6,7,8))
    for idx, letter in zip(range(3,9), ["C","D","E","F","G","H"]):
        ws2.cell(row=last2+1, column=idx).value = f"=SUM({letter}3:{letter}{last2})"
        ws2.cell(row=last2+1, column=idx).number_format = "₱#,##0.00"
    ws2.cell(row=last2+1, column=9).value = f"=G{last2+1}/H{last2+1}"
    ws2.cell(row=last2+1, column=9).number_format = "0.00"
    _autofit(ws2)

    # Sheet 3 — Ad Efficiency
    ws3 = wb.create_sheet("3. Ad Efficiency")
    _write_headers(ws3, "Ad Channel Efficiency",
                   ["Channel", "Week Spend", "% of Budget", "ROAS", "Rank"])
    ch_data = sorted([
        (n, float(wk[sc].sum()), float(wk[rc].sum()))
        for n, sc, rc in [("Meta","meta_ads","shopify"),("Shopee Ads","shopee_ads","shopee"),
                           ("Lazada Ads","lazada_ads","lazada"),("TikTok Ads","tiktok_ads","tiktok"),
                           ("Google Ads","google_ads","shopify"),("Shopee CPAs","shopee_cpas","shopee"),
                           ("Lazada CPAs","lazada_cpas","lazada")]
    ], key=lambda x: x[1]/x[2] if x[2] else 0, reverse=True)
    total_wk_spend = float(wk["total_ads_spent"].sum())
    for rank, (name, spend, rev) in enumerate(ch_data, 1):
        i = rank + 2
        roas = round(rev/spend, 2) if spend > 0 else 0
        ws3.cell(row=i, column=1, value=name)
        ws3.cell(row=i, column=2, value=spend)
        ws3.cell(row=i, column=3, value=spend/total_wk_spend if total_wk_spend else 0)
        ws3.cell(row=i, column=4, value=roas)
        ws3.cell(row=i, column=5, value=rank)
        _style(ws3, i, 5, i % 2 == 0, currency=(2,), pct=(3,), roas=(4,))
    _autofit(ws3)

    # Sheet 4 — Campaign Impact
    ws4 = wb.create_sheet("4. Campaign Impact")
    _write_headers(ws4, "Campaign Impact This Week",
                   ["Date", "Campaign", "Revenue", "Avg Normal Day", "Lift ₱", "Lift %"])
    norm = float(df[df["remarks"] == ""]["total_revenue"].mean())
    camps = wk[wk["remarks"] != ""]
    if camps.empty:
        ws4.cell(row=3, column=1, value="No campaigns flagged this week.")
    else:
        for i, (_, r) in enumerate(camps.iterrows(), 3):
            rev = float(r["total_revenue"])
            ws4.cell(row=i, column=1, value=r["date"].strftime("%Y-%m-%d"))
            ws4.cell(row=i, column=2, value=str(r["remarks"]))
            ws4.cell(row=i, column=3, value=rev)
            ws4.cell(row=i, column=4, value=norm)
            ws4.cell(row=i, column=5, value=rev - norm)
            ws4.cell(row=i, column=6, value=_pdelta(rev, norm))
            _style(ws4, i, 6, i % 2 == 0, currency=(3,4,5), pct=(6,))
    _autofit(ws4)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────
#  REPORT 3 — MONTHLY
# ─────────────────────────────────────────────────────────────

def generate_monthly_report(df: pd.DataFrame) -> io.BytesIO | None:
    monthly = (
        df.assign(period=df["date"].dt.to_period("M"))
        .groupby("period")
        .agg(gross=("total_revenue","sum"), ex_vat=("total_revenue_ex_vat","sum"),
             shopify=("shopify","sum"), shopee=("shopee","sum"),
             lazada=("lazada","sum"), tiktok=("tiktok","sum"),
             ads=("total_ads_spent","sum"), meta=("meta_ads","sum"),
             shopee_ads=("shopee_ads","sum"), lazada_ads=("lazada_ads","sum"),
             tiktok_ads=("tiktok_ads","sum"), google=("google_ads","sum"))
        .reset_index()
    )
    monthly["roas"]    = monthly.apply(lambda r: round(r["gross"]/r["ads"],2) if r["ads"] else 0, axis=1)
    monthly["mom"]     = monthly["gross"].pct_change()
    monthly["mth_str"] = monthly["period"].dt.strftime("%b %Y")

    wb = Workbook()

    # Sheet 1 — Summary
    ws1 = wb.active
    ws1.title = "1. Monthly Summary"
    _write_headers(ws1, "Monthly P&L Summary — Sola Body (Gloss & Lace Inc.)",
                   ["Month", "Gross Revenue", "Revenue ex-VAT", "Total Ad Spend", "ROAS", "MoM Growth"])
    for i, (_, r) in enumerate(monthly.iterrows(), 3):
        ws1.cell(row=i, column=1, value=r["mth_str"])
        ws1.cell(row=i, column=2, value=float(r["gross"]))
        ws1.cell(row=i, column=3, value=float(r["ex_vat"]))
        ws1.cell(row=i, column=4, value=float(r["ads"]))
        ws1.cell(row=i, column=5, value=r["roas"])
        ws1.cell(row=i, column=6, value=r["mom"])
        _style(ws1, i, 6, i % 2 == 0, currency=(2,3,4), roas=(5,), pct=(6,))
    last1 = len(monthly) + 2
    _total(ws1, last1+1, ["GRAND TOTAL"]+[None]*5, currency=(2,3,4))
    for idx, l in zip([2,3,4], ["B","C","D"]):
        ws1.cell(row=last1+1, column=idx).value = f"=SUM({l}3:{l}{last1})"
        ws1.cell(row=last1+1, column=idx).number_format = "₱#,##0.00"
    ws1.cell(row=last1+1, column=5).value = f"=B{last1+1}/D{last1+1}"
    ws1.cell(row=last1+1, column=5).number_format = "0.00"
    _autofit(ws1)

    # Sheet 2 — Platform MoM
    ws2 = wb.create_sheet("2. Platform MoM")
    _write_headers(ws2, "Platform Revenue — Month over Month",
                   ["Month","Shopify","Shopee","Lazada","TikTok","Total","Shopify %","Shopee %","Lazada %","TikTok %"])
    for i, (_, r) in enumerate(monthly.iterrows(), 3):
        vals = [r["mth_str"], float(r["shopify"]), float(r["shopee"]),
                float(r["lazada"]), float(r["tiktok"]), float(r["gross"])]
        for j, v in enumerate(vals, 1):
            ws2.cell(row=i, column=j, value=v)
        for j, col in zip([7,8,9,10], ["B","C","D","E"]):
            ws2.cell(row=i, column=j, value=f"={col}{i}/F{i}")
        _style(ws2, i, 10, i % 2 == 0, currency=(2,3,4,5,6), pct=(7,8,9,10))
    _autofit(ws2)

    # Sheet 3 — Ads & ROAS
    ws3 = wb.create_sheet("3. Ads & ROAS")
    _write_headers(ws3, "Monthly Ad Spend & ROAS",
                   ["Month","Meta","Shopee Ads","Lazada Ads","TikTok Ads","Google","Total Spend","ROAS","Spend % of Rev"])
    for i, (_, r) in enumerate(monthly.iterrows(), 3):
        vals = [r["mth_str"], float(r["meta"]), float(r["shopee_ads"]),
                float(r["lazada_ads"]), float(r["tiktok_ads"]), float(r["google"]),
                float(r["ads"]), r["roas"],
                round(float(r["ads"])/float(r["gross"]), 4) if r["gross"] else 0]
        for j, v in enumerate(vals, 1):
            ws3.cell(row=i, column=j, value=v)
        _style(ws3, i, 9, i % 2 == 0, currency=(2,3,4,5,6,7), roas=(8,), pct=(9,))
    _autofit(ws3)

    # Sheet 4 — Growth Tracker
    ws4 = wb.create_sheet("4. Growth Tracker")
    _write_headers(ws4, "Revenue Growth Tracker",
                   ["Month", "Monthly Revenue", "MoM Growth", "Cumulative Revenue"])
    for i, (_, r) in enumerate(monthly.iterrows(), 3):
        ws4.cell(row=i, column=1, value=r["mth_str"])
        ws4.cell(row=i, column=2, value=float(r["gross"]))
        ws4.cell(row=i, column=3, value=r["mom"])
        ws4.cell(row=i, column=4, value=f"=SUM($B$3:B{i})")
        _style(ws4, i, 4, i % 2 == 0, currency=(2, 4), pct=(3,))
    _autofit(ws4)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────
#  REPORT 4 — ADS AUDIT
# ─────────────────────────────────────────────────────────────

def generate_ads_audit(df: pd.DataFrame) -> io.BytesIO:
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "1. ROAS by Channel"
    period = f"{df['date'].min().strftime('%b %d')}–{df['date'].max().strftime('%b %d, %Y')}"
    _write_headers(ws1, f"Ad Spend Audit — {period}",
                   ["Rank", "Channel", "Total Spend", "Revenue", "ROAS", "% of Budget", "Status"])
    channels = [("Meta Ads","meta_ads","shopify"),("Shopee Ads","shopee_ads","shopee"),
                ("Lazada Ads","lazada_ads","lazada"),("TikTok Ads","tiktok_ads","tiktok"),
                ("Google Ads","google_ads","shopify"),("Shopee CPAs","shopee_cpas","shopee"),
                ("Lazada CPAs","lazada_cpas","lazada")]
    total_spend = float(df["total_ads_spent"].sum())
    results = sorted(
        [(n, float(df[sc].sum()), float(df[rc].sum())) for n, sc, rc in channels],
        key=lambda x: x[1]/x[2] if x[2] else 0, reverse=True
    )
    for rank, (name, spend, rev) in enumerate(results, 1):
        i     = rank + 2
        roas  = round(rev/spend, 2) if spend > 0 else 0
        share = spend/total_spend if total_spend else 0
        status = "Good ✓" if roas >= 4 else ("Watch ⚠" if roas >= 2 else "Review ✗")
        for j, v in enumerate([rank, name, spend, rev, roas, share, status], 1):
            ws1.cell(row=i, column=j, value=v)
        _style(ws1, i, 7, i % 2 == 0, currency=(3,4), pct=(6,), roas=(5,))
        color = "006400" if "Good" in status else ("CC6600" if "Watch" in status else "CC0000")
        ws1.cell(row=i, column=7).font = Font(name="Arial", size=10, bold=True, color=color)
    _autofit(ws1)

    ws2 = wb.create_sheet("2. Daily Spend Log")
    _write_headers(ws2, "Daily Ad Spend Log",
                   ["Date","Day","Meta","Shopee Ads","Lazada Ads","TikTok Ads",
                    "Google","Shopee CPAs","Lazada CPAs","Total Spend","Total Revenue","ROAS","Campaign"])
    for i, (_, r) in enumerate(df.iterrows(), 3):
        vals = [r["date"].strftime("%Y-%m-%d"), r["day_of_week"],
                float(r["meta_ads"]), float(r["shopee_ads"]), float(r["lazada_ads"]),
                float(r["tiktok_ads"]), float(r["google_ads"]),
                float(r["shopee_cpas"]), float(r["lazada_cpas"]),
                float(r["total_ads_spent"]), float(r["total_revenue"]),
                round(float(r["roas"]), 2), str(r.get("remarks") or "")]
        for j, v in enumerate(vals, 1):
            ws2.cell(row=i, column=j, value=v)
        _style(ws2, i, 13, i % 2 == 0, currency=(3,4,5,6,7,8,9,10,11), roas=(12,))
        if r.get("remarks"):
            ws2.cell(row=i, column=13).font = Font(name="Arial", size=10, bold=True, color="185FA5")
    _autofit(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────
#  FLASK ROUTES
# ─────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/upload", methods=["POST"])
def upload():
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400
    file = request.files["file"]
    if not file.filename:
        return jsonify({"error": "Empty filename"}), 400
    try:
        df  = load_and_clean(file)
        sid = str(uuid.uuid4())[:12]
        DATA_STORE[sid] = df
        return jsonify({
            "session_id": sid,
            "rows":       len(df),
            "start":      df["date"].min().strftime("%Y-%m-%d"),
            "end":        df["date"].max().strftime("%Y-%m-%d"),
            "dates":      df["date"].dt.strftime("%Y-%m-%d").tolist(),
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/daily")
def api_daily():
    sid      = request.args.get("session_id")
    date_str = request.args.get("date")
    df       = DATA_STORE.get(sid)
    if df is None:
        return jsonify({"error": "Session expired. Please re-upload your file."}), 404
    sel  = pd.Timestamp(date_str)
    row  = df[df["date"] == sel]
    if row.empty:
        return jsonify({"error": "No data for this date"}), 404
    t    = row.iloc[0]
    y    = df[df["date"] == sel - pd.Timedelta(days=1)]
    lw   = df[df["date"] == sel - pd.Timedelta(days=7)]
    mtd  = df[(df["date"].dt.month == sel.month) & (df["date"].dt.year == sel.year) & (df["date"] <= sel)]
    mtdr = float(mtd["total_revenue"].sum())
    mtda = float(mtd["total_ads_spent"].sum())
    return jsonify({
        "date":       date_str,
        "day":        t["day_of_week"],
        "remarks":    str(t.get("remarks") or ""),
        "kpis": {
            "total_revenue":       round(float(t["total_revenue"]), 2),
            "total_revenue_ex_vat":round(float(t["total_revenue_ex_vat"]), 2),
            "total_ads_spent":     round(float(t["total_ads_spent"]), 2),
            "roas":                round(float(t["roas"]), 2),
            "mtd_revenue":         round(mtdr, 2),
            "mtd_ads":             round(mtda, 2),
            "mtd_roas":            round(mtdr / mtda, 2) if mtda > 0 else 0,
        },
        "deltas": {
            "rev_vs_yesterday": _pct(float(t["total_revenue"]), y, "total_revenue"),
            "rev_vs_lastweek":  _pct(float(t["total_revenue"]), lw, "total_revenue"),
            "ads_vs_yesterday": _pct(float(t["total_ads_spent"]), y, "total_ads_spent"),
        },
        "platforms": {
            "shopify": round(float(t["shopify"]), 2),
            "shopee":  round(float(t["shopee"]),  2),
            "lazada":  round(float(t["lazada"]),  2),
            "tiktok":  round(float(t["tiktok"]),  2),
        },
        "ads": {
            "Meta Ads":    round(float(t["meta_ads"]),    2),
            "Shopee Ads":  round(float(t["shopee_ads"]),  2),
            "Lazada Ads":  round(float(t["lazada_ads"]),  2),
            "TikTok Ads":  round(float(t["tiktok_ads"]),  2),
            "Google Ads":  round(float(t["google_ads"]),  2),
            "Shopee CPAs": round(float(t["shopee_cpas"]), 2),
            "Lazada CPAs": round(float(t["lazada_cpas"]), 2),
        },
    })


@app.route("/api/trend")
def api_trend():
    sid      = request.args.get("session_id")
    date_str = request.args.get("date")
    days     = int(request.args.get("days", 30))
    df       = DATA_STORE.get(sid)
    if df is None:
        return jsonify({"error": "Session expired"}), 404
    sel   = pd.Timestamp(date_str)
    trend = df[df["date"] <= sel].tail(days)
    return jsonify({
        "dates":    trend["date"].dt.strftime("%Y-%m-%d").tolist(),
        "revenue":  [round(float(x), 2) for x in trend["total_revenue"]],
        "ads":      [round(float(x), 2) for x in trend["total_ads_spent"]],
        "roas":     [round(float(x), 2) for x in trend["roas"]],
        "shopify":  [round(float(x), 2) for x in trend["shopify"]],
        "shopee":   [round(float(x), 2) for x in trend["shopee"]],
        "lazada":   [round(float(x), 2) for x in trend["lazada"]],
        "tiktok":   [round(float(x), 2) for x in trend["tiktok"]],
        "campaigns":[{"date": r["date"].strftime("%Y-%m-%d"), "label": r["remarks"]}
                     for _, r in trend.iterrows() if r.get("remarks")],
    })


@app.route("/download/<report_type>")
def download_report(report_type):
    sid      = request.args.get("session_id")
    date_str = request.args.get("date", "")
    df       = DATA_STORE.get(sid)
    if df is None:
        return jsonify({"error": "Session expired. Re-upload your file."}), 404

    buf, filename = None, ""
    if report_type == "daily" and date_str:
        buf      = generate_daily_report(df, pd.Timestamp(date_str))
        filename = f"Sola_Daily_{date_str}.xlsx"
    elif report_type == "weekly":
        buf      = generate_weekly_report(df)
        filename = f"Sola_Weekly_W{int(df['date'].max().isocalendar().week)}.xlsx"
    elif report_type == "monthly":
        buf      = generate_monthly_report(df)
        filename = f"Sola_Monthly_{df['date'].max().strftime('%Y-%m')}.xlsx"
    elif report_type == "audit":
        buf      = generate_ads_audit(df)
        filename = f"Sola_AdsAudit_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    else:
        return jsonify({"error": "Unknown report type"}), 400

    if buf is None:
        return jsonify({"error": "Could not generate report"}), 500

    return send_file(
        buf,
        download_name=filename,
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    print("\n  Sola Body Revenue Command Center")
    print("  Open your browser at: http://localhost:5000\n")
    app.run(debug=True, port=5000)
