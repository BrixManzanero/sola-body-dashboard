"""
Sola Body — Daily Revenue Command Center
Streamlit Cloud version
Deploy: push this file + requirements.txt to GitHub root → connect to Streamlit Cloud
"""

import io
import warnings
from datetime import datetime, timedelta

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ── Page config ───────────────────────────────────────────────
st.set_page_config(
    page_title="Sola Body — Revenue Command Center",
    page_icon="📊",
    layout="wide",
)

# ── Constants ─────────────────────────────────────────────────
EXCEL_ORIGIN  = datetime(1899, 12, 30)
PLATFORM_COLS = ["shopify", "shopee", "lazada", "tiktok"]
AD_COLS       = ["meta_ads", "shopee_ads", "lazada_ads",
                 "tiktok_ads", "google_ads", "shopee_cpas", "lazada_cpas"]
COL_MAP       = {
    0: "dow", 1: "month", 2: "date_raw",
    3: "shopify", 4: "shopee", 5: "lazada", 6: "tiktok",
    7: "total_revenue", 8: "total_revenue_ex_vat",
    9: "meta_ads", 10: "shopee_ads", 11: "lazada_ads",
    12: "tiktok_ads", 13: "google_ads",
    14: "shopee_cpas", 15: "lazada_cpas",
    20: "total_ads_spent",
}


# ─────────────────────────────────────────────────────────────
#  DATA LOADING & CLEANING
# ─────────────────────────────────────────────────────────────

def _to_float(val) -> float:
    if pd.isna(val):
        return 0.0
    if isinstance(val, (int, float)):
        return 0.0 if (isinstance(val, float) and np.isnan(val)) else float(val)
    s = str(val).replace("₱","").replace(",","").replace(" ","").replace("-","").strip()
    try:
        return float(s) if s else 0.0
    except ValueError:
        return 0.0


def _parse_date(val):
    if pd.isna(val) or str(val).strip() in ("", "nan"):
        return None
    try:
        p = pd.to_datetime(val)
        if pd.notna(p):
            return p
    except Exception:
        pass
    try:
        n = float(val)
        if 40000 < n < 60000:
            return pd.Timestamp(EXCEL_ORIGIN + timedelta(days=int(n)))
    except Exception:
        pass
    return None


@st.cache_data(show_spinner="Reading and cleaning your data…")
def load_and_clean(file_bytes: bytes) -> pd.DataFrame:
    """Load, clean and cache the BR Input Excel."""
    buf = io.BytesIO(file_bytes)
    xl  = pd.ExcelFile(buf)

    sheet = next(
        (s for s in xl.sheet_names if "daily" in s.lower()),
        xl.sheet_names[0],
    )
    raw = pd.read_excel(buf, sheet_name=sheet, header=None, dtype=str)

    header_idx = 0
    for i, row in raw.iterrows():
        vals = [str(v).strip().lower() for v in row if pd.notna(v) and str(v).strip()]
        if "date" in vals or "shopify" in vals:
            header_idx = i
            break

    data = raw.iloc[header_idx + 3:].copy().reset_index(drop=True)
    data.columns = [COL_MAP.get(i, f"col_{i}") for i in range(len(data.columns))]

    data["date"] = pd.to_datetime(
        data["date_raw"].apply(_parse_date), errors="coerce"
    )
    df = data[data["date"].notna()].copy()

    num_cols = PLATFORM_COLS + ["total_revenue", "total_revenue_ex_vat"] + AD_COLS + ["total_ads_spent"]
    for col in num_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col].apply(_to_float), errors="coerce").fillna(0.0)

    rev   = df["total_revenue"].astype(float)
    spend = df["total_ads_spent"].astype(float)
    df["roas"]        = np.where(spend > 0, rev / spend, 0.0)
    df["week_number"] = df["date"].dt.isocalendar().week.astype(int)
    df["month_name"]  = df["date"].dt.strftime("%b %Y")
    df["day_of_week"] = df["date"].dt.day_name()
    df["remarks"]     = df["dow"].apply(
        lambda v: str(v).strip()
        if pd.notna(v) and len(str(v).strip()) > 3
        and str(v).strip().upper()
        not in {"SUN","MON","TUE","WED","THU","FRI","SAT","NAN"}
        else ""
    )
    return df.sort_values("date").reset_index(drop=True)


# ─────────────────────────────────────────────────────────────
#  EXCEL REPORT HELPERS
# ─────────────────────────────────────────────────────────────

DARK  = "1D3557"
MID   = "457B9D"
LIGHT = "A8DADC"
ALT   = "F1FAEE"
THIN  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


def _fill(c): return PatternFill("solid", fgColor=c)
def _font(c="FFFFFF", s=10, b=True): return Font(name="Arial", bold=b, color=c, size=s)

def _hdr(ws, title, cols, row=1):
    n = len(cols)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n)
    c = ws.cell(row=row, column=1, value=title)
    c.font = _font(s=12); c.fill = _fill(DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 22
    for j, name in enumerate(cols, 1):
        c = ws.cell(row=row+1, column=j, value=name)
        c.font = _font(s=10); c.fill = _fill(MID)
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = THIN
    ws.row_dimensions[row+1].height = 28

def _autofit(ws, mn=10, mx=30):
    for col in ws.columns:
        l = get_column_letter(col[0].column)
        ws.column_dimensions[l].width = max(mn, min(mx, max(len(str(c.value or "")) for c in col)+2))

def _style(ws, row, n, alt, cur=(), pct=(), roas=()):
    fill = _fill(ALT) if alt else None
    for j in range(1, n+1):
        c = ws.cell(row=row, column=j)
        c.border = THIN; c.font = Font(name="Arial", size=10)
        if fill: c.fill = fill
        if j in cur: c.number_format = "₱#,##0.00"
        if j in pct and isinstance(c.value, float):
            c.number_format = "0.0%"
            c.font = Font(name="Arial", size=10, bold=True,
                          color="006400" if c.value >= 0 else "CC0000")
        if j in roas and isinstance(c.value, (int, float)):
            c.number_format = "0.00"
            color = "006400" if c.value >= 4 else ("CC6600" if c.value >= 2 else "CC0000")
            c.font = Font(name="Arial", size=10, bold=True, color=color)

def _total(ws, row, vals, cur=()):
    for j, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=j, value=v)
        c.fill = _fill(LIGHT); c.font = Font(name="Arial", size=10, bold=True); c.border = THIN
        if j in cur: c.number_format = "₱#,##0.00"

def _pd(a, b):
    try:
        if b and b != 0: return (a - b) / abs(b)
    except Exception: pass
    return None


# ─────────────────────────────────────────────────────────────
#  REPORT GENERATORS
# ─────────────────────────────────────────────────────────────

def gen_daily(df, sel):
    row = df[df["date"] == sel]
    if row.empty: return None
    t   = row.iloc[0]
    y   = df[df["date"] == sel - pd.Timedelta(days=1)]
    lw  = df[df["date"] == sel - pd.Timedelta(days=7)]
    mtd = df[(df["date"].dt.month == sel.month) & (df["date"].dt.year == sel.year) & (df["date"] <= sel)]
    mtdr = float(mtd["total_revenue"].sum())
    mtda = float(mtd["total_ads_spent"].sum())

    wb = Workbook()
    ws1 = wb.active; ws1.title = "1. Snapshot"
    _hdr(ws1, f"Daily Snapshot — {sel.strftime('%B %d, %Y')}", ["Metric","Value","vs Yesterday","vs Last Week"])
    yrev = float(y.iloc[0]["total_revenue"])  if not y.empty  else None
    lwrev= float(lw.iloc[0]["total_revenue"]) if not lw.empty else None
    rows = [
        ("Total Revenue",        float(t["total_revenue"]),        _pd(float(t["total_revenue"]),   yrev), _pd(float(t["total_revenue"]), lwrev)),
        ("Total Revenue ex-VAT", float(t["total_revenue_ex_vat"]), None, None),
        ("Total Ad Spend",       float(t["total_ads_spent"]),      _pd(float(t["total_ads_spent"]), float(y.iloc[0]["total_ads_spent"]) if not y.empty else None), None),
        ("ROAS",                 round(float(t["roas"]),2),        None, None),
        ("MTD Revenue",          mtdr, None, None),
        ("MTD Ad Spend",         mtda, None, None),
        ("MTD ROAS",             round(mtdr/mtda,2) if mtda>0 else 0, None, None),
        ("Campaign / Event",     str(t.get("remarks") or "—"),     None, None),
    ]
    for i,(label,val,d1,d2) in enumerate(rows, 3):
        ws1.cell(row=i,column=1,value=label); ws1.cell(row=i,column=2,value=val)
        ws1.cell(row=i,column=3,value=d1);   ws1.cell(row=i,column=4,value=d2)
        is_cur = label not in ("ROAS","MTD ROAS","Campaign / Event")
        _style(ws1,i,4,i%2==0, cur=(2,) if is_cur else (), pct=(3,4))
        if "ROAS" in label: ws1.cell(row=i,column=2).number_format="0.00"
    _autofit(ws1)

    ws2 = wb.create_sheet("2. Platform Revenue")
    _hdr(ws2, f"Platform Revenue — {sel.strftime('%B %d, %Y')}", ["Platform","Revenue","% of Total","vs Yesterday","vs Last Week"])
    total = float(t["total_revenue"])
    for i,(name,col) in enumerate([("Shopify","shopify"),("Shopee","shopee"),("Lazada","lazada"),("TikTok Shop","tiktok")],3):
        val=float(t[col]); share=val/total if total>0 else 0
        ws2.cell(row=i,column=1,value=name); ws2.cell(row=i,column=2,value=val)
        ws2.cell(row=i,column=3,value=share)
        ws2.cell(row=i,column=4,value=_pd(val, float(y.iloc[0][col]) if not y.empty else None))
        ws2.cell(row=i,column=5,value=_pd(val, float(lw.iloc[0][col]) if not lw.empty else None))
        _style(ws2,i,5,i%2==0,cur=(2,),pct=(3,4,5))
    last2=4+2; _total(ws2,last2+1,["TOTAL",None,None,None,None],cur=(2,))
    ws2.cell(row=last2+1,column=2).value=f"=SUM(B3:B{last2})"
    ws2.cell(row=last2+1,column=2).number_format="₱#,##0.00"
    _autofit(ws2)

    ws3 = wb.create_sheet("3. Ad Spend")
    _hdr(ws3, f"Ad Spend — {sel.strftime('%B %d, %Y')}", ["Channel","Spend","% of Budget","Platform Revenue","ROAS"])
    chs=[("Meta Ads","meta_ads","shopify"),("Shopee Ads","shopee_ads","shopee"),
         ("Lazada Ads","lazada_ads","lazada"),("TikTok Ads","tiktok_ads","tiktok"),
         ("Google Ads","google_ads","shopify"),("Shopee CPAs","shopee_cpas","shopee"),
         ("Lazada CPAs","lazada_cpas","lazada")]
    ts=float(t["total_ads_spent"])
    for i,(name,sc,rc) in enumerate(chs,3):
        sp=float(t[sc]); rv=float(t[rc])
        ws3.cell(row=i,column=1,value=name); ws3.cell(row=i,column=2,value=sp)
        ws3.cell(row=i,column=3,value=sp/ts if ts>0 else 0); ws3.cell(row=i,column=4,value=rv)
        ws3.cell(row=i,column=5,value=round(rv/sp,2) if sp>0 else 0)
        _style(ws3,i,5,i%2==0,cur=(2,4),pct=(3,),roas=(5,))
    last3=len(chs)+2; _total(ws3,last3+1,["TOTAL",None,None,None,None],cur=(2,4))
    ws3.cell(row=last3+1,column=2).value=f"=SUM(B3:B{last3})"
    ws3.cell(row=last3+1,column=2).number_format="₱#,##0.00"
    _autofit(ws3)

    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf


def gen_weekly(df):
    last=df["date"].max(); wn=int(last.isocalendar().week); yr=last.year
    wk=df[(df["week_number"]==wn)&(df["date"].dt.year==yr)]
    prev=df[(df["week_number"]==wn-1)&(df["date"].dt.year==yr)]
    if wk.empty: return None
    ws=wk["date"].min().strftime("%b %d"); we=wk["date"].max().strftime("%b %d, %Y")
    wb=Workbook()

    ws1=wb.active; ws1.title="1. Weekly Summary"
    _hdr(ws1,f"Weekly Summary — {ws} to {we}",["Metric","This Week","Last Week","WoW Change"])
    tr=float(wk["total_revenue"].sum()); pr=float(prev["total_revenue"].sum()) if not prev.empty else 0
    ts=float(wk["total_ads_spent"].sum()); ps=float(prev["total_ads_spent"].sum()) if not prev.empty else 0
    pw=lambda c: float(prev[c].sum()) if not prev.empty else 0
    sumrows=[
        ("Total Revenue",tr,pr,_pd(tr,pr)),
        ("Shopify",float(wk["shopify"].sum()),pw("shopify"),_pd(float(wk["shopify"].sum()),pw("shopify"))),
        ("Shopee", float(wk["shopee"].sum()), pw("shopee"), _pd(float(wk["shopee"].sum()), pw("shopee"))),
        ("Lazada", float(wk["lazada"].sum()), pw("lazada"), _pd(float(wk["lazada"].sum()), pw("lazada"))),
        ("TikTok Shop",float(wk["tiktok"].sum()),pw("tiktok"),_pd(float(wk["tiktok"].sum()),pw("tiktok"))),
        ("Total Ad Spend",ts,ps,_pd(ts,ps)),
        ("ROAS",round(tr/ts,2) if ts else 0, round(pr/ps,2) if ps else 0, None),
        ("Best Day", wk.loc[wk["total_revenue"].idxmax(),"day_of_week"],"—",None),
        ("Worst Day",wk.loc[wk["total_revenue"].idxmin(),"day_of_week"],"—",None),
    ]
    for i,(label,tw,pv,wow) in enumerate(sumrows,3):
        ws1.cell(row=i,column=1,value=label); ws1.cell(row=i,column=2,value=tw)
        ws1.cell(row=i,column=3,value=pv);   ws1.cell(row=i,column=4,value=wow)
        ic=label not in("ROAS","Best Day","Worst Day")
        _style(ws1,i,4,i%2==0,cur=(2,3) if ic else (),pct=(4,))
        if label=="ROAS": ws1.cell(row=i,column=2).number_format="0.00"; ws1.cell(row=i,column=3).number_format="0.00"
    _autofit(ws1)

    ws2=wb.create_sheet("2. Daily Breakdown")
    _hdr(ws2,f"Daily Breakdown — Week {wn}",["Date","Day","Shopify","Shopee","Lazada","TikTok","Total Revenue","Ad Spend","ROAS","Campaign"])
    for i,(_,r) in enumerate(wk.iterrows(),3):
        for j,v in enumerate([r["date"].strftime("%Y-%m-%d"),r["day_of_week"],
            float(r["shopify"]),float(r["shopee"]),float(r["lazada"]),float(r["tiktok"]),
            float(r["total_revenue"]),float(r["total_ads_spent"]),round(float(r["roas"]),2),
            str(r.get("remarks") or "")],1): ws2.cell(row=i,column=j,value=v)
        _style(ws2,i,10,i%2==0,cur=(3,4,5,6,7,8),roas=(9,))
    last2=len(wk)+2; _total(ws2,last2+1,["TOTAL",""]+[None]*8,cur=(3,4,5,6,7,8))
    for idx,l in zip(range(3,9),["C","D","E","F","G","H"]):
        ws2.cell(row=last2+1,column=idx).value=f"=SUM({l}3:{l}{last2})"
        ws2.cell(row=last2+1,column=idx).number_format="₱#,##0.00"
    _autofit(ws2)

    ws3=wb.create_sheet("3. Ad Efficiency")
    _hdr(ws3,"Ad Channel Efficiency",["Channel","Week Spend","% of Budget","ROAS","Rank"])
    chd=sorted([(n,float(wk[sc].sum()),float(wk[rc].sum()))
        for n,sc,rc in [("Meta","meta_ads","shopify"),("Shopee Ads","shopee_ads","shopee"),
        ("Lazada Ads","lazada_ads","lazada"),("TikTok Ads","tiktok_ads","tiktok"),
        ("Google Ads","google_ads","shopify"),("Shopee CPAs","shopee_cpas","shopee"),
        ("Lazada CPAs","lazada_cpas","lazada")]],key=lambda x:x[1]/x[2] if x[2] else 0,reverse=True)
    tws=float(wk["total_ads_spent"].sum())
    for rank,(name,sp,rv) in enumerate(chd,1):
        i=rank+2; roas=round(rv/sp,2) if sp>0 else 0
        ws3.cell(row=i,column=1,value=name); ws3.cell(row=i,column=2,value=sp)
        ws3.cell(row=i,column=3,value=sp/tws if tws else 0); ws3.cell(row=i,column=4,value=roas); ws3.cell(row=i,column=5,value=rank)
        _style(ws3,i,5,i%2==0,cur=(2,),pct=(3,),roas=(4,))
    _autofit(ws3)

    ws4=wb.create_sheet("4. Campaign Impact")
    _hdr(ws4,"Campaign Impact",["Date","Campaign","Revenue","Avg Normal Day","Lift ₱","Lift %"])
    norm=float(df[df["remarks"]==""]['total_revenue'].mean())
    camps=wk[wk["remarks"]!=""]
    if camps.empty: ws4.cell(row=3,column=1,value="No campaigns flagged this week.")
    else:
        for i,(_,r) in enumerate(camps.iterrows(),3):
            rv=float(r["total_revenue"])
            for j,v in enumerate([r["date"].strftime("%Y-%m-%d"),str(r["remarks"]),rv,norm,rv-norm,_pd(rv,norm)],1):
                ws4.cell(row=i,column=j,value=v)
            _style(ws4,i,6,i%2==0,cur=(3,4,5),pct=(6,))
    _autofit(ws4)

    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf


def gen_monthly(df):
    m=(df.assign(period=df["date"].dt.to_period("M"))
        .groupby("period").agg(gross=("total_revenue","sum"),ex_vat=("total_revenue_ex_vat","sum"),
        shopify=("shopify","sum"),shopee=("shopee","sum"),lazada=("lazada","sum"),tiktok=("tiktok","sum"),
        ads=("total_ads_spent","sum"),meta=("meta_ads","sum"),sa=("shopee_ads","sum"),
        la=("lazada_ads","sum"),ta=("tiktok_ads","sum"),ga=("google_ads","sum")).reset_index())
    m["roas"]=m.apply(lambda r:round(r["gross"]/r["ads"],2) if r["ads"] else 0,axis=1)
    m["mom"]=m["gross"].pct_change(); m["ms"]=m["period"].dt.strftime("%b %Y")
    wb=Workbook()

    ws1=wb.active; ws1.title="1. Monthly Summary"
    _hdr(ws1,"Monthly P&L Summary — Sola Body (Gloss & Lace Inc.)",["Month","Gross Revenue","Revenue ex-VAT","Total Ad Spend","ROAS","MoM Growth"])
    for i,(_,r) in enumerate(m.iterrows(),3):
        for j,v in enumerate([r["ms"],float(r["gross"]),float(r["ex_vat"]),float(r["ads"]),r["roas"],r["mom"]],1):
            ws1.cell(row=i,column=j,value=v)
        _style(ws1,i,6,i%2==0,cur=(2,3,4),roas=(5,),pct=(6,))
    last1=len(m)+2; _total(ws1,last1+1,["GRAND TOTAL"]+[None]*5,cur=(2,3,4))
    for idx,l in zip([2,3,4],["B","C","D"]):
        ws1.cell(row=last1+1,column=idx).value=f"=SUM({l}3:{l}{last1})"
        ws1.cell(row=last1+1,column=idx).number_format="₱#,##0.00"
    ws1.cell(row=last1+1,column=5).value=f"=B{last1+1}/D{last1+1}"; ws1.cell(row=last1+1,column=5).number_format="0.00"
    _autofit(ws1)

    ws2=wb.create_sheet("2. Platform MoM")
    _hdr(ws2,"Platform Revenue — MoM",["Month","Shopify","Shopee","Lazada","TikTok","Total","Shopify %","Shopee %","Lazada %","TikTok %"])
    for i,(_,r) in enumerate(m.iterrows(),3):
        for j,v in enumerate([r["ms"],float(r["shopify"]),float(r["shopee"]),float(r["lazada"]),float(r["tiktok"]),float(r["gross"])],1):
            ws2.cell(row=i,column=j,value=v)
        for j,col in zip([7,8,9,10],["B","C","D","E"]): ws2.cell(row=i,column=j,value=f"={col}{i}/F{i}")
        _style(ws2,i,10,i%2==0,cur=(2,3,4,5,6),pct=(7,8,9,10))
    _autofit(ws2)

    ws3=wb.create_sheet("3. Ads & ROAS")
    _hdr(ws3,"Monthly Ad Spend & ROAS",["Month","Meta","Shopee Ads","Lazada Ads","TikTok Ads","Google","Total Spend","ROAS","Spend % of Rev"])
    for i,(_,r) in enumerate(m.iterrows(),3):
        for j,v in enumerate([r["ms"],float(r["meta"]),float(r["sa"]),float(r["la"]),float(r["ta"]),float(r["ga"]),
            float(r["ads"]),r["roas"],round(float(r["ads"])/float(r["gross"]),4) if r["gross"] else 0],1):
            ws3.cell(row=i,column=j,value=v)
        _style(ws3,i,9,i%2==0,cur=(2,3,4,5,6,7),roas=(8,),pct=(9,))
    _autofit(ws3)

    ws4=wb.create_sheet("4. Growth Tracker")
    _hdr(ws4,"Revenue Growth Tracker",["Month","Monthly Revenue","MoM Growth","Cumulative Revenue"])
    for i,(_,r) in enumerate(m.iterrows(),3):
        ws4.cell(row=i,column=1,value=r["ms"]); ws4.cell(row=i,column=2,value=float(r["gross"]))
        ws4.cell(row=i,column=3,value=r["mom"]); ws4.cell(row=i,column=4,value=f"=SUM($B$3:B{i})")
        _style(ws4,i,4,i%2==0,cur=(2,4),pct=(3,))
    _autofit(ws4)

    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf


def gen_audit(df):
    wb=Workbook(); ws1=wb.active; ws1.title="1. ROAS by Channel"
    period=f"{df['date'].min().strftime('%b %d')}–{df['date'].max().strftime('%b %d, %Y')}"
    _hdr(ws1,f"Ad Spend Audit — {period}",["Rank","Channel","Total Spend","Revenue","ROAS","% of Budget","Status"])
    chs=[("Meta Ads","meta_ads","shopify"),("Shopee Ads","shopee_ads","shopee"),
         ("Lazada Ads","lazada_ads","lazada"),("TikTok Ads","tiktok_ads","tiktok"),
         ("Google Ads","google_ads","shopify"),("Shopee CPAs","shopee_cpas","shopee"),
         ("Lazada CPAs","lazada_cpas","lazada")]
    ts=float(df["total_ads_spent"].sum())
    res=sorted([(n,float(df[sc].sum()),float(df[rc].sum())) for n,sc,rc in chs],
               key=lambda x:x[1]/x[2] if x[2] else 0,reverse=True)
    for rank,(name,sp,rv) in enumerate(res,1):
        i=rank+2; roas=round(rv/sp,2) if sp>0 else 0
        status="Good ✓" if roas>=4 else ("Watch ⚠" if roas>=2 else "Review ✗")
        for j,v in enumerate([rank,name,sp,rv,roas,sp/ts if ts else 0,status],1):
            ws1.cell(row=i,column=j,value=v)
        _style(ws1,i,7,i%2==0,cur=(3,4),pct=(6,),roas=(5,))
        color="006400" if "Good" in status else("CC6600" if "Watch" in status else "CC0000")
        ws1.cell(row=i,column=7).font=Font(name="Arial",size=10,bold=True,color=color)
    _autofit(ws1)

    ws2=wb.create_sheet("2. Daily Spend Log")
    _hdr(ws2,"Daily Ad Spend Log",["Date","Day","Meta","Shopee Ads","Lazada Ads","TikTok Ads",
         "Google","Shopee CPAs","Lazada CPAs","Total Spend","Total Revenue","ROAS","Campaign"])
    for i,(_,r) in enumerate(df.iterrows(),3):
        for j,v in enumerate([r["date"].strftime("%Y-%m-%d"),r["day_of_week"],
            float(r["meta_ads"]),float(r["shopee_ads"]),float(r["lazada_ads"]),
            float(r["tiktok_ads"]),float(r["google_ads"]),float(r["shopee_cpas"]),
            float(r["lazada_cpas"]),float(r["total_ads_spent"]),float(r["total_revenue"]),
            round(float(r["roas"]),2),str(r.get("remarks") or "")],1):
            ws2.cell(row=i,column=j,value=v)
        _style(ws2,i,13,i%2==0,cur=(3,4,5,6,7,8,9,10,11),roas=(12,))
    _autofit(ws2)
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf


# ─────────────────────────────────────────────────────────────
#  STREAMLIT UI
# ─────────────────────────────────────────────────────────────

st.title("📊 Sola Body — Revenue Command Center")
st.caption("Upload your BR Input Excel to view your dashboard and download reports.")

# ── Sidebar ───────────────────────────────────────────────────
with st.sidebar:
    st.header("Controls")
    uploaded = st.file_uploader(
        "Upload BR Input Excel",
        type=["xlsx", "xls"],
        help="Requires sheet: Daily Revenue and Ads Expense!",
    )

if uploaded is None:
    st.info("👈 Upload your BR Input Excel file from the sidebar to get started.")
    st.markdown("""
    **Expected file:** Your daily revenue and ads Excel (e.g. `BR_Input.xlsx`)  
    **Required sheet:** `Daily Revenue and Ads Expense!`
    """)
    st.stop()

# ── Load data ─────────────────────────────────────────────────
file_bytes = uploaded.read()
df = load_and_clean(file_bytes)

if df.empty:
    st.error("No valid data found. Make sure the sheet is named 'Daily Revenue and Ads Expense!'")
    st.stop()

st.success(f"✅ {len(df):,} days loaded — {df['date'].min().strftime('%b %d, %Y')} to {df['date'].max().strftime('%b %d, %Y')}")

# ── Sidebar controls ──────────────────────────────────────────
with st.sidebar:
    dates   = sorted(df["date"].dt.date.unique(), reverse=True)
    sel_date= st.selectbox("View date", dates,
                           format_func=lambda d: pd.Timestamp(d).strftime("%b %d, %Y (%a)"))
    trend_d = st.slider("Trend — days to show", 7, 90, 30)
    st.divider()
    st.markdown("**Download Reports**")

sel_dt   = pd.Timestamp(sel_date)
t_row    = df[df["date"] == sel_dt]
if t_row.empty:
    st.warning("No data for selected date.")
    st.stop()

t        = t_row.iloc[0]
yest_row = df[df["date"] == sel_dt - pd.Timedelta(days=1)]
lw_row   = df[df["date"] == sel_dt - pd.Timedelta(days=7)]
mtd_df   = df[(df["date"].dt.month == sel_dt.month) & (df["date"].dt.year == sel_dt.year) & (df["date"] <= sel_dt)]

def _delta(new, old_df, col):
    if old_df.empty: return None
    old = float(old_df.iloc[0][col])
    return (new - old) / abs(old) if old != 0 else None

# ── Day header ────────────────────────────────────────────────
st.subheader(f"📅 {sel_dt.strftime('%A, %B %d, %Y')}")
if t["remarks"]:
    st.info(f"📌 Campaign: **{t['remarks']}**")

# ── KPI Cards ─────────────────────────────────────────────────
k1, k2, k3, k4, k5, k6 = st.columns(6)
k1.metric("Total Revenue",    f"₱{float(t['total_revenue']):,.0f}",
          f"{_delta(float(t['total_revenue']),yest_row,'total_revenue'):.1%}" if _delta(float(t['total_revenue']),yest_row,'total_revenue') is not None else None)
k2.metric("Revenue ex-VAT",   f"₱{float(t['total_revenue_ex_vat']):,.0f}")
k3.metric("Ad Spend",         f"₱{float(t['total_ads_spent']):,.0f}",
          f"{_delta(float(t['total_ads_spent']),yest_row,'total_ads_spent'):.1%}" if _delta(float(t['total_ads_spent']),yest_row,'total_ads_spent') is not None else None,
          delta_color="inverse")
k4.metric("ROAS",             f"{float(t['roas']):.2f}x")
k5.metric("MTD Revenue",      f"₱{float(mtd_df['total_revenue'].sum()):,.0f}")
k6.metric("MTD ROAS",         f"{float(mtd_df['total_revenue'].sum())/float(mtd_df['total_ads_spent'].sum()):.2f}x"
          if float(mtd_df['total_ads_spent'].sum()) > 0 else "—")

st.divider()

# ── Platform + Ads charts ─────────────────────────────────────
col_l, col_r = st.columns(2)
with col_l:
    st.markdown("**Revenue by Platform**")
    plat_df = pd.DataFrame({
        "Platform": ["Shopify","Shopee","Lazada","TikTok"],
        "Revenue":  [float(t["shopify"]),float(t["shopee"]),float(t["lazada"]),float(t["tiktok"])],
    }).sort_values("Revenue", ascending=True)
    fig1 = px.bar(plat_df, x="Revenue", y="Platform", orientation="h",
                  color="Platform", text_auto=".2s",
                  color_discrete_map={"Shopify":"#5C832F","Shopee":"#EE4D2D","Lazada":"#0F146D","TikTok":"#555555"})
    fig1.update_layout(showlegend=False, height=230, margin=dict(l=0,r=0,t=0,b=0))
    fig1.update_xaxes(tickprefix="₱", tickformat=",.0f")
    st.plotly_chart(fig1, use_container_width=True)

with col_r:
    st.markdown("**Ad Spend by Channel**")
    ad_df = pd.DataFrame({
        "Channel": ["Meta","Shopee Ads","Lazada Ads","TikTok Ads","Google","Shopee CPAs","Lazada CPAs"],
        "Spend":   [float(t["meta_ads"]),float(t["shopee_ads"]),float(t["lazada_ads"]),
                    float(t["tiktok_ads"]),float(t["google_ads"]),float(t["shopee_cpas"]),float(t["lazada_cpas"])],
    })
    ad_df = ad_df[ad_df["Spend"] > 0].sort_values("Spend", ascending=True)
    fig2 = px.bar(ad_df, x="Spend", y="Channel", orientation="h",
                  color_discrete_sequence=["#457B9D"], text_auto=".2s")
    fig2.update_layout(showlegend=False, height=230, margin=dict(l=0,r=0,t=0,b=0))
    fig2.update_xaxes(tickprefix="₱", tickformat=",.0f")
    st.plotly_chart(fig2, use_container_width=True)

st.divider()

# ── Revenue trend ──────────────────────────────────────────────
st.markdown(f"**Revenue Trend — Last {trend_d} Days**")
trend_df = df[df["date"] <= sel_dt].tail(trend_d).copy()
fig3 = go.Figure()
fig3.add_trace(go.Scatter(x=trend_df["date"], y=trend_df["total_revenue"],
    name="Revenue", line=dict(color="#1D3557", width=2),
    fill="tozeroy", fillcolor="rgba(29,53,87,0.08)"))
fig3.add_trace(go.Scatter(x=trend_df["date"], y=trend_df["total_ads_spent"],
    name="Ad Spend", line=dict(color="#E63946", width=1.5, dash="dash")))
camps = trend_df[trend_df["remarks"] != ""]
if not camps.empty:
    fig3.add_trace(go.Scatter(x=camps["date"], y=camps["total_revenue"], mode="markers",
        name="Campaign", marker=dict(color="#F59E0B", size=10, symbol="star")))
fig3.update_layout(height=300, margin=dict(l=0,r=0,t=10,b=0),
    yaxis=dict(tickprefix="₱", tickformat=",.0f"),
    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))
st.plotly_chart(fig3, use_container_width=True)

# ── Platform area chart ───────────────────────────────────────
st.markdown("**Platform Share Over Time**")
fig4 = px.area(trend_df, x="date", y=["shopify","shopee","lazada","tiktok"],
    color_discrete_map={"shopify":"#5C832F","shopee":"#EE4D2D","lazada":"#0F146D","tiktok":"#555555"},
    labels={"value":"Revenue (₱)","variable":"Platform"})
fig4.update_layout(height=260, margin=dict(l=0,r=0,t=10,b=0),
    yaxis=dict(tickprefix="₱", tickformat=",.0f"),
    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))
st.plotly_chart(fig4, use_container_width=True)

st.divider()

# ── Download Reports ──────────────────────────────────────────
st.markdown("### 📥 Download Excel Reports")
d1, d2, d3, d4 = st.columns(4)

with d1:
    buf = gen_daily(df, sel_dt)
    if buf:
        st.download_button("📄 Daily Report", data=buf,
            file_name=f"Sola_Daily_{sel_dt.strftime('%Y-%m-%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)
with d2:
    buf = gen_weekly(df)
    if buf:
        st.download_button("📅 Weekly Report", data=buf,
            file_name=f"Sola_Weekly_W{int(df['date'].max().isocalendar().week)}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)
with d3:
    buf = gen_monthly(df)
    if buf:
        st.download_button("📆 Monthly Report", data=buf,
            file_name=f"Sola_Monthly_{df['date'].max().strftime('%Y-%m')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)
with d4:
    buf = gen_audit(df)
    if buf:
        st.download_button("📊 Ads Audit", data=buf,
            file_name=f"Sola_AdsAudit_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)

# ── Sidebar downloads (same buttons, sidebar position) ────────
with st.sidebar:
    buf = gen_daily(df, sel_dt)
    if buf:
        st.download_button("📄 Daily Report", data=buf,
            file_name=f"Sola_Daily_{sel_dt.strftime('%Y-%m-%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)
    buf = gen_weekly(df)
    if buf:
        st.download_button("📅 Weekly Report", data=buf,
            file_name=f"Sola_Weekly_W{int(df['date'].max().isocalendar().week)}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)
    buf = gen_monthly(df)
    if buf:
        st.download_button("📆 Monthly Report", data=buf,
            file_name=f"Sola_Monthly_{df['date'].max().strftime('%Y-%m')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)
    buf = gen_audit(df)
    if buf:
        st.download_button("📊 Ads Audit", data=buf,
            file_name=f"Sola_AdsAudit_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)

st.caption("Sola Body Revenue Command Center · Gloss and Lace Inc.")
